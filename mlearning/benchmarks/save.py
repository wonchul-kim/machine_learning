def save_df_by_class_to_pdf(performance_by_class, output_filename, idx2class=None):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.units import inch

    # PDF 생성
    document = SimpleDocTemplate(output_filename, pagesize=letter)
    styles = getSampleStyleSheet()

    # 데이터 직접 PDF에 추가
    content = []

    # 제목 추가
    content.append(Paragraph('Performances', styles['Title']))

    # 데이터 추가
    for data in performance_by_class:
        for key, value in data.items():
            if isinstance(value, list):
                value_str = ', '.join(map(str, value))
            else:
                if key == 'class': 
                    value_str = str(idx2class[value])
                else:
                    value_str = str(value)
            content.append(Paragraph(f'{key}: {value_str}', styles['BodyText']))

        content.append(Paragraph('<br/><br/>', styles['Normal']))  # 추가적인 여백

    # PDF 생성
    document.build(content)


def save_pf_by_image_to_excel(performance_by_image, output_filename, idx2class=None):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment
    
    # 데이터를 정리하여 리스트로 변환
    rows = []

    for filename, classes in performance_by_image.items():
        for class_id, metrics in classes.items():
            row = {'filename': filename, 'class': class_id if idx2class is None else idx2class[class_id]}
            row.update(metrics)
            rows.append(row)

    # 리스트를 DataFrame으로 변환
    df = pd.DataFrame(rows)

    # DataFrame을 엑셀 파일로 저장
    df.to_excel(output_filename, index=False, engine='openpyxl')


    # 엑셀 파일 로드
    workbook = load_workbook(output_filename)
    sheet = workbook.active

    # 현재 시트에서 'filename' 열 병합 수행
    filename_col = 'A'  # 'filename' 열이 A열에 있다고 가정

    # 병합할 시작 위치와 종료 위치 추적
    start_row = 2  # 데이터는 2행부터 시작하므로 2행부터 확인
    end_row = start_row

    fill_color = PatternFill(start_color="D9DDDC", end_color="D9DDDC", fill_type="solid")
    header_fill = PatternFill(start_color="D9DDDC", end_color="D9DDDC", fill_type="solid")  # 연한 파란색

    # 테두리 스타일 지정 (굵은 선)
    thick_border_rb = Border(right=Side(style='thick'), bottom=Side(style='thick'))
    thick_border_rt = Border(right=Side(style='thick'), top=Side(style='thick'))
    thick_border_r = Border(right=Side(style='thick'))
    thick_border_rtb = Border(right=Side(style='thick'), bottom=Side(style='thick'), top=Side(style='thick'))

    # 첫 번째 행(헤더)에 색상 적용
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.border = thick_border_rb

    # 행을 순회하면서 동일한 filename에 대해 병합 작업 수행
    for row in range(2, sheet.max_row + 1):
        sheet[f'A{row}'].fill = fill_color

        current_value = sheet[f'{filename_col}{row}'].value
        next_value = sheet[f'{filename_col}{row + 1}'].value if row + 1 <= sheet.max_row else None
        
        if current_value == next_value:
            end_row += 1
        else:
            if start_row != end_row:
                # 셀 병합
                sheet.merge_cells(f'{filename_col}{start_row}:{filename_col}{end_row}')
                
                # 병합된 범위에 대해 위아래 테두리 굵게 설정
                for merged_row in range(start_row, end_row + 1):
                    sheet[f'A{merged_row}'].border = thick_border_rb
                    
                for col in range(1, sheet.max_column + 1):
                    col_letter = sheet.cell(row=1, column=col).column_letter
                    for merged_row in range(start_row, end_row + 1):
                        if merged_row == start_row:
                            sheet[f'{col_letter}{merged_row}'].border = thick_border_rt
                        elif merged_row == end_row:
                            sheet[f'{col_letter}{merged_row}'].border = thick_border_rb
                        else:
                            sheet[f'{col_letter}{merged_row}'].border = thick_border_r
                    
            else:
                # 병합이 없는 경우에도 테두리와 색상 적용
                sheet[f'A{start_row}'].fill = fill_color
                sheet[f'A{start_row}'].border = thick_border_rtb
                
                for col in range(1, sheet.max_column + 1):
                    col_letter = sheet.cell(row=1, column=col).column_letter
                    sheet[f'{col_letter}{start_row}'].border = thick_border_rtb
                    
            start_row = row + 1
            end_row = start_row


    # 가운데 정렬 설정
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 모든 셀에 대해 가운데 정렬 적용
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = center_alignment


    # 엑셀 파일 저장
    workbook.save(output_filename)
    
    
if __name__ == '__main__':
    
    # 데이터 정의 (예제 데이터)
    performance_by_class = {
        'class': 'Class A',
        'accumulated_precision': [0.8, 0.85, 0.9],
        'accumulated_recall': [0.7, 0.75, 0.8],
        'precision': 0.9,
        'recall': 0.8,
        'ap': 0.85,
        'interpolated_precision': [0.8, 0.85, 0.9],
        'interpolated_recall': [0.7, 0.75, 0.8],
        'total_gt': 100,
        'total_tp': 80,
        'total_fp': 10,
        'total_fn': 20,
    }
    output_filename = '/HDD/datasets/projects/sungjin/body/benchmark/pf_by_class.pdf'
    
    save_df_by_class_to_pdf(performance_by_class, output_filename)